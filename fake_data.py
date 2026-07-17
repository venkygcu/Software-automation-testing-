from faker import Faker
from openpyxl import Workbook

# Create workbook and worksheet
wb = Workbook()
ws = wb.active

# Create Faker object
fake_data = Faker()

# Generate fake data
for i in range(1, 11):
   ws.cell(row=i, column=1).value = fake_data.name()
   ws.cell(row=i, column=2).value = fake_data.email()

# Save the workbook
wb.save("fake_data.xlsx")

print("Excel file created successfully!")